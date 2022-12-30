
import os
import csv

####### DELETE FILE ######
def delete_view():
    folder_path = os.environ["USERPROFILE"]+'\Downloads'
    for file in os.listdir(folder_path):
        if(file.startswith("view")):
            filepath  = os.path.join(folder_path,file)
            os.remove(filepath)




######## PARSE THE FILE #######
def collate_data():
    folder_path = os.environ["USERPROFILE"]+'\Downloads'
    dict_of_views = {}

    def create_dict(key,value,cs_name):
        dict_of_views[key] = {
            "value" : value,
            "cs_name" : cs_name
        }

    def check_row(data,header_object):
        
        if (len(data) == 0):
                return False
        #if(data.count("incontinence_protector_size") > 0) or (data.count("^incontinence_protector_size$") > 0):
                #check for value and customer 
        #if(header.find("value") and header.("customer_name")):
        if header_object['value'] != None and header_object['customer_name'] != None:
                    value =  data[header_object["value"]]
                    cs_name = data[header_object["customer_name"]]
                    item_id = data[header_object["item_id"]]
                    create_dict(item_id,value,cs_name)
                    return True
                    
        return False
            
                    



    for file in os.listdir(folder_path) :
        if file.startswith("view"):
            f = os.path.join(folder_path,file)
            print(f)
            with open(f, 'r') as file:
                csvreader = csv.reader(file)
                header_object = {}
                header = []
                body  = []
                index = 0 
                for row in csvreader:
                    if(index == 0):
                        header = row
                    else:
                        body.append(row)
                    index+=1
                print(header)
                for ind,head_value in enumerate(header):
                    if head_value.endswith("value"):
                        header_object["value"] = ind
                    elif head_value.endswith("customer_name"):
                        header_object["customer_name"] = ind
                    elif head_value.endswith("attribute"):
                        header_object["attribute"] = ind
                    elif head_value.endswith("item_id"):
                        header_object["item_id"] = ind
                print(header_object)
                

                if check_row(body[0],header_object) == False:
                        print(header_object["item_id"] , "item")
                        create_dict(body[0][header_object["item_id"]],"No","No")


    print(dict_of_views)





### Read and Write to file

    import openpyxl
    folder_path = 'input.xlsx'

    workbook = openpyxl.load_workbook(filename=folder_path)
    sheet = workbook.active

    for row in range(2,sheet.max_row+1):
        asin = sheet.cell(row = row, column = 1)
        view_data = dict_of_views.get(asin.value)
        #print(asin.value)
        if view_data:
            for col in range(3,5):
                cell_obj = sheet.cell(row = row, column = col)
                cell_obj.value = view_data.get("value") if col == 3 else view_data.get("cs_name")
    workbook.save(filename=folder_path)


